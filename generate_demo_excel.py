"""
Run this script once to generate clients_data.xlsx for the dashboard demo.
Requires: pip install openpyxl
"""

import openpyxl
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
wb.remove(wb.active)

today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
d0 = today.strftime("%Y-%m-%d")
d1 = (today + timedelta(days=1)).strftime("%Y-%m-%d")
d2 = (today + timedelta(days=2)).strftime("%Y-%m-%d")

CLIENTS = [
    {
        "sheet_name": "Flour Pastificio",
        "meta": {
            "CLIENT_NAME": "Flour Pastificio",
            "ADDRESS": "Via della Croce 18, Roma",
            "AGENT_NAME": "Marco",
            "AGENT_STATUS": "Online",
        },
        "stats": {
            "calls_today": 24,
            "reservations": 8,
            "conversion_pct": 33,
            "covers_booked": 29,
            "escalations": 3,
        },
        "allergy_alerts": [
            {"name": "Valentina Conti",   "time": "20:00", "phone": "+39 06 7733 2211", "allergy": "Nut allergy, severe"},
            {"name": "Alessandro Greco",  "time": "20:30", "phone": "+39 349 223 4418", "allergy": "Dairy intolerant x2"},
        ],
        "escalations": [
            {"name": "Unknown Caller",  "phone": "+39 06 3344 5521", "type": "Private Hire",    "note": "Enquired about exclusive private hire for 35 guests. Pricing and availability require owner decision before callback.", "received": "11:42"},
            {"name": "Giorgio Mancini", "phone": "+39 349 771 8832", "type": "Complaint",        "note": "Caller unhappy with wait time on last visit and requested a manager callback to discuss compensation.", "received": "14:15"},
            {"name": "Unknown Caller",  "phone": "+39 06 5593 2284", "type": "Language Barrier", "note": "Caller spoke no Italian or English — possible French speaker. Marco could not complete booking.", "received": "16:03"},
        ],
        "reservations": [
            {"name": "Martina Ferraro",    "date": d0, "time": "12:30", "guests": 2, "allergy": "",                    "phone": "+39 06 4521 8834"},
            {"name": "Davide Romano",      "date": d0, "time": "13:00", "guests": 4, "allergy": "Gluten intolerant",    "phone": "+39 338 742 1093"},
            {"name": "Elena Ricci",        "date": d0, "time": "13:30", "guests": 2, "allergy": "",                    "phone": "+39 06 9921 4455"},
            {"name": "Marco Salvatore",    "date": d0, "time": "19:30", "guests": 3, "allergy": "",                    "phone": "+39 347 882 3310"},
            {"name": "Valentina Conti",    "date": d0, "time": "20:00", "guests": 2, "allergy": "Nut allergy, severe",  "phone": "+39 06 7733 2211"},
            {"name": "Alessandro Greco",   "date": d0, "time": "20:30", "guests": 6, "allergy": "Dairy intolerant x2", "phone": "+39 349 223 4418"},
            {"name": "Sofia De Luca",      "date": d0, "time": "21:00", "guests": 2, "allergy": "",                    "phone": "+39 333 551 9920"},
            {"name": "Federico Marchetti", "date": d1, "time": "13:00", "guests": 4, "allergy": "",                    "phone": "+39 06 3312 7741"},
            {"name": "Giulia Bianchi",     "date": d1, "time": "20:00", "guests": 2, "allergy": "Shellfish allergy",   "phone": "+39 348 991 2206"},
            {"name": "Luca Esposito",      "date": d1, "time": "20:30", "guests": 5, "allergy": "",                    "phone": "+39 339 441 5572"},
            {"name": "Isabella Fontana",   "date": d1, "time": "21:00", "guests": 3, "allergy": "Gluten-free required","phone": "+39 06 4411 3390"},
            {"name": "Chiara Lombardi",    "date": d2, "time": "19:30", "guests": 4, "allergy": "",                    "phone": "+39 345 770 3318"},
            {"name": "Antonio Moretti",    "date": d2, "time": "20:00", "guests": 2, "allergy": "",                    "phone": "+39 06 6612 8843"},
            {"name": "Roberto Caruso",     "date": d2, "time": "20:30", "guests": 6, "allergy": "Tree nut allergy",    "phone": "+39 06 8831 2290"},
        ],
        "weekly_calls": [44, 32, 16, 19, 21, 26, 24],
        "weekly_res":   [18, 12,  6,  8,  8, 11,  8],
    },
    {
        "sheet_name": "The Harbour Grill",
        "meta": {
            "CLIENT_NAME": "The Harbour Grill",
            "ADDRESS": "12 Harbour Road, Dun Laoghaire, Dublin",
            "AGENT_NAME": "Aoife",
            "AGENT_STATUS": "Online",
        },
        "stats": {
            "calls_today": 18,
            "reservations": 6,
            "conversion_pct": 44,
            "covers_booked": 22,
            "escalations": 1,
        },
        "allergy_alerts": [
            {"name": "Siobhan Murphy", "time": "19:00", "phone": "+353 87 234 5678", "allergy": "Shellfish allergy"},
        ],
        "escalations": [
            {"name": "Patrick Walsh", "phone": "+353 86 987 6543", "type": "Large Group", "note": "Caller enquired about reserving the full terrace for a 40-person birthday dinner. Manager approval required.", "received": "13:05"},
        ],
        "reservations": [
            {"name": "Siobhan Murphy",    "date": d0, "time": "19:00", "guests": 2, "allergy": "Shellfish allergy", "phone": "+353 87 234 5678"},
            {"name": "Conor O'Brien",     "date": d0, "time": "19:30", "guests": 4, "allergy": "",                  "phone": "+353 85 321 9876"},
            {"name": "Aoife Byrne",       "date": d0, "time": "20:00", "guests": 3, "allergy": "",                  "phone": "+353 83 456 7890"},
            {"name": "Niamh Kelly",       "date": d0, "time": "20:30", "guests": 2, "allergy": "",                  "phone": "+353 87 765 4321"},
            {"name": "Declan Fitzpatrick","date": d1, "time": "13:00", "guests": 6, "allergy": "",                  "phone": "+353 86 543 2109"},
            {"name": "Fiona Lynch",       "date": d1, "time": "19:00", "guests": 2, "allergy": "Vegan",             "phone": "+353 85 678 9012"},
            {"name": "Sean Doyle",        "date": d2, "time": "19:30", "guests": 5, "allergy": "",                  "phone": "+353 87 890 1234"},
        ],
        "weekly_calls": [28, 22, 14, 18, 16, 20, 18],
        "weekly_res":   [12,  9,  5,  7,  6,  8,  6],
    },
    {
        "sheet_name": "La Brasserie",
        "meta": {
            "CLIENT_NAME": "La Brasserie",
            "ADDRESS": "45 Rue du Faubourg Saint-Antoine, Paris 11e",
            "AGENT_NAME": "Claire",
            "AGENT_STATUS": "Online",
        },
        "stats": {
            "calls_today": 31,
            "reservations": 14,
            "conversion_pct": 45,
            "covers_booked": 48,
            "escalations": 2,
        },
        "allergy_alerts": [
            {"name": "Marie Dubois",    "time": "20:00", "phone": "+33 6 12 34 56 78", "allergy": "Peanut allergy"},
            {"name": "Pierre Martin",   "time": "20:45", "phone": "+33 6 98 76 54 32", "allergy": "Coeliac disease"},
            {"name": "Sophie Lefevre",  "time": "21:15", "phone": "+33 6 55 44 33 22", "allergy": "Lactose intolerant"},
        ],
        "escalations": [
            {"name": "Jacques Renard",  "phone": "+33 1 44 55 66 77", "type": "Complaint",    "note": "Guest dissatisfied with previous reservation mix-up. Requested manager call and complimentary dessert on next visit.", "received": "10:20"},
            {"name": "Unknown Caller",  "phone": "+33 6 77 88 99 00", "type": "Press Inquiry", "note": "Caller identified as food journalist requesting an interview and tasting menu booking. Needs owner approval.", "received": "15:47"},
        ],
        "reservations": [
            {"name": "Marie Dubois",     "date": d0, "time": "12:30", "guests": 2, "allergy": "Peanut allergy",    "phone": "+33 6 12 34 56 78"},
            {"name": "Antoine Bernard",  "date": d0, "time": "13:00", "guests": 4, "allergy": "",                  "phone": "+33 6 23 45 67 89"},
            {"name": "Camille Morel",    "date": d0, "time": "13:30", "guests": 2, "allergy": "",                  "phone": "+33 6 34 56 78 90"},
            {"name": "Isabelle Petit",   "date": d0, "time": "19:30", "guests": 5, "allergy": "",                  "phone": "+33 6 45 67 89 01"},
            {"name": "Pierre Martin",    "date": d0, "time": "20:45", "guests": 2, "allergy": "Coeliac disease",   "phone": "+33 6 98 76 54 32"},
            {"name": "Sophie Lefevre",   "date": d0, "time": "21:15", "guests": 3, "allergy": "Lactose intolerant","phone": "+33 6 55 44 33 22"},
            {"name": "Francois Girard",  "date": d1, "time": "20:00", "guests": 4, "allergy": "",                  "phone": "+33 6 56 78 90 12"},
            {"name": "Amelie Fournier",  "date": d1, "time": "20:30", "guests": 2, "allergy": "Nut allergy",       "phone": "+33 6 67 89 01 23"},
            {"name": "Thomas Rousseau",  "date": d1, "time": "21:00", "guests": 6, "allergy": "",                  "phone": "+33 6 78 90 12 34"},
            {"name": "Margot Vincent",   "date": d2, "time": "12:30", "guests": 2, "allergy": "",                  "phone": "+33 6 89 01 23 45"},
            {"name": "Nicolas Laurent",  "date": d2, "time": "19:30", "guests": 3, "allergy": "",                  "phone": "+33 6 90 12 34 56"},
            {"name": "Elise Simon",      "date": d2, "time": "20:00", "guests": 4, "allergy": "Vegan",             "phone": "+33 6 01 23 45 67"},
        ],
        "weekly_calls": [52, 41, 28, 35, 38, 44, 31],
        "weekly_res":   [23, 18, 11, 14, 15, 19, 14],
    },
]


def write_client_sheet(wb, client):
    ws = wb.create_sheet(title=client["sheet_name"])
    r = 1

    # ── META ──
    for key, val in client["meta"].items():
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # ── STATS ──
    ws.cell(row=r, column=1, value="SECTION")
    ws.cell(row=r, column=2, value="STATS")
    r += 1
    stats = client["stats"]
    keys = list(stats.keys())
    for i, k in enumerate(keys, 1):
        ws.cell(row=r, column=i, value=k)
    r += 1
    for i, k in enumerate(keys, 1):
        ws.cell(row=r, column=i, value=stats[k])
    r += 2

    # ── ALLERGY_ALERTS ──
    ws.cell(row=r, column=1, value="SECTION")
    ws.cell(row=r, column=2, value="ALLERGY_ALERTS")
    r += 1
    for i, h in enumerate(["name", "time", "phone", "allergy"], 1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for a in client["allergy_alerts"]:
        ws.cell(row=r, column=1, value=a["name"])
        ws.cell(row=r, column=2, value=a["time"])
        ws.cell(row=r, column=3, value=a["phone"])
        ws.cell(row=r, column=4, value=a["allergy"])
        r += 1
    r += 1

    # ── ESCALATIONS ──
    ws.cell(row=r, column=1, value="SECTION")
    ws.cell(row=r, column=2, value="ESCALATIONS")
    r += 1
    for i, h in enumerate(["name", "phone", "type", "note", "received"], 1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for e in client["escalations"]:
        ws.cell(row=r, column=1, value=e["name"])
        ws.cell(row=r, column=2, value=e["phone"])
        ws.cell(row=r, column=3, value=e["type"])
        ws.cell(row=r, column=4, value=e["note"])
        ws.cell(row=r, column=5, value=e["received"])
        r += 1
    r += 1

    # ── RESERVATIONS ──
    ws.cell(row=r, column=1, value="SECTION")
    ws.cell(row=r, column=2, value="RESERVATIONS")
    r += 1
    for i, h in enumerate(["name", "date", "time", "guests", "allergy", "phone"], 1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for res in client["reservations"]:
        ws.cell(row=r, column=1, value=res["name"])
        # Store date as plain text to avoid Excel auto-conversion
        c = ws.cell(row=r, column=2, value=res["date"])
        c.number_format = "@"
        ws.cell(row=r, column=3, value=res["time"])
        ws.cell(row=r, column=4, value=res["guests"])
        ws.cell(row=r, column=5, value=res["allergy"])
        ws.cell(row=r, column=6, value=res["phone"])
        r += 1
    r += 1

    # ── WEEKLY ──
    ws.cell(row=r, column=1, value="SECTION")
    ws.cell(row=r, column=2, value="WEEKLY")
    r += 1
    for i, h in enumerate(["day", "calls", "reservations"], 1):
        ws.cell(row=r, column=i, value=h)
    r += 1
    for i in range(7):
        ws.cell(row=r, column=1, value=f"Day {i + 1}")
        ws.cell(row=r, column=2, value=client["weekly_calls"][i])
        ws.cell(row=r, column=3, value=client["weekly_res"][i])
        r += 1


for client in CLIENTS:
    write_client_sheet(wb, client)

wb.save("clients_data.xlsx")
print(f"Created clients_data.xlsx with {len(CLIENTS)} client tabs:")
for c in CLIENTS:
    print(f"  - {c['sheet_name']}")
